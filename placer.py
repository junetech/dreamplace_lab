import logging
import time
from pathlib import PurePath
from typing import Dict

import numpy as np
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import dreamplace.configure as configure
from Params import Params
from PlaceDB import PlaceDB


def make_area_numpin_count_xlsx(placedb: PlaceDB, aux_input: str):
    # number of pins for each node
    num_pins_in_nodes = np.zeros(placedb.num_nodes)
    for node_id in range(placedb.num_physical_nodes):
        num_pins_in_nodes[node_id] = len(placedb.node2pin_map[node_id])

    node_size_x = np.array(placedb.node_size_x).astype(np.int32)
    node_size_y = np.array(placedb.node_size_y).astype(np.int32)

    # node area -> number of pins -> number of nodes
    movable_count_dict: Dict[float, Dict[int, int]] = {}

    # iterate over movable nodes
    for node_id in range(placedb.num_movable_nodes):
        node_area = node_size_x[node_id] * node_size_y[node_id]
        if node_area not in movable_count_dict:
            movable_count_dict[node_area] = {}
        num_pins = num_pins_in_nodes[node_id]
        if num_pins not in movable_count_dict[node_area]:
            movable_count_dict[node_area][num_pins] = 1
        else:
            movable_count_dict[node_area][num_pins] += 1

    # node area -> number of pins -> number of nodes
    fixed_count_dict: Dict[float, Dict[int, int]] = {}

    for node_id in range(
        placedb.num_movable_nodes, placedb.num_movable_nodes + placedb.num_terminals
    ):
        node_area = node_size_x[node_id] * node_size_y[node_id]
        if node_area not in fixed_count_dict:
            fixed_count_dict[node_area] = {}
        num_pins = num_pins_in_nodes[node_id]
        if num_pins not in fixed_count_dict[node_area]:
            fixed_count_dict[node_area][num_pins] = 1
        else:
            fixed_count_dict[node_area][num_pins] += 1

    # write to workbook
    wb = Workbook()

    # movable sheet
    mv_ws: Worksheet = wb.active
    mv_ws.title = "Movable"
    # fixed sheet
    fx_ws: Worksheet = wb.create_sheet("Fixed")
    col_header = ["NodeArea", "#pins", "Count"]
    mv_ws.append(col_header)
    fx_ws.append(col_header)

    for node_area, num_pins_dict in movable_count_dict.items():
        for num_pins, count in num_pins_dict.items():
            mv_ws.append([node_area, num_pins, count])

    for node_area, num_pins_dict in fixed_count_dict.items():
        for num_pins, count in num_pins_dict.items():
            fx_ws.append([node_area, num_pins, count])

    xlsx_filename = PurePath(aux_input).stem + ".xlsx"
    # save the workbook
    wb.save(filename=xlsx_filename)
    logging.info(f"Saved node area & #pins count to {xlsx_filename}")


def make_numpin_count_xlsx(placedb: PlaceDB, aux_input: str):
    # number of pins for each node
    num_pins_in_nodes = np.array(
        [
            len(placedb.node2pin_map[node_id])
            for node_id in range(placedb.num_physical_nodes)
        ]
    )
    # number of pins for each edge
    num_pins_in_edges = np.array(
        [len(placedb.net2pin_map[net_id]) for net_id in range(len(placedb.net2pin_map))]
    )
    logging.info("Total number of pins is %d" % np.sum(num_pins_in_nodes))

    # write to workbook
    wb = Workbook()

    # node sheet
    node_ws: Worksheet = wb.active
    node_ws.title = "node"
    # edge sheet
    edge_ws: Worksheet = wb.create_sheet("edge")
    col_header = ["Degree", "#pins"]
    node_ws.append(col_header)
    edge_ws.append(col_header)

    for count, degree in zip(
        *np.histogram(num_pins_in_nodes, bins=sorted(set(num_pins_in_nodes)))
    ):
        node_ws.append([degree, count])

    for count, degree in zip(
        *np.histogram(num_pins_in_edges, bins=sorted(set(num_pins_in_edges)))
    ):
        edge_ws.append([degree, count])

    xlsx_filename = PurePath(aux_input).stem + "_pincount.xlsx"
    # save the workbook
    wb.save(filename=xlsx_filename)
    logging.info(f"Saved node-wise & edge-wise pin count to {xlsx_filename}")


def place(params: Params):
    # assert (not params.gpu) or configure.compile_configurations[
    #     "CUDA_FOUND"
    # ] == "TRUE", "CANNOT enable GPU without CUDA compiled"

    np.random.seed(params.random_seed)
    # read database
    tt = time.time()
    placedb = PlaceDB()
    placedb(params)
    # make_area_numpin_count_xlsx(placedb, params.aux_input)
    make_numpin_count_xlsx(placedb, params.aux_input)
    proc_time = time.time() - tt
    logging.info(f"Process: Input takes {proc_time:.3f} sec")
