import ast
import datetime
import logging
import re
import sys
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

START_DT: datetime.datetime

PLACER_LOG_FILENAME = "dreamplace_lab.log"
PARAMS_PATTERN = r"^\[INFO\]root-parameters=(\{.*\})"
TIME_PATTERN = r"^\[INFO\]root-Process:(.*)takes(\d+\.\d+)sec$"
OBJ_PATTERN = r"^\[INFO\]root-Process:(.*)haswHPWLof(\d+\.\d+)&overflowof(\d+\.\d+)"
OUTPUT_FILENAME = "dreamplace_runs.xlsx"


def parse_a_line(a_line: str) -> Dict[str, Dict[str, Any]]:
    parsed_dict = {}
    key_str = ""

    _a_line = a_line.replace(" ", "")
    params_regex = re.compile(PARAMS_PATTERN)
    mo = params_regex.search(_a_line)
    if mo:
        key_str = "Input"
        params_dict = ast.literal_eval(mo.group(1))
        if params_dict["def_input"]:
            input_path = params_dict["def_input"]
        elif params_dict["aux_input"]:
            input_path = params_dict["aux_input"]
        parsed_dict[key_str] = {"path": input_path}
    else:
        time_regex = re.compile(TIME_PATTERN)
        mo = time_regex.search(_a_line)
        if mo:
            key_str, runtime = mo.groups()
            parsed_dict[key_str] = {"runtime": runtime}
        else:
            obj_regex = re.compile(OBJ_PATTERN)
            mo = obj_regex.search(_a_line)
            if mo:
                key_str, w_hpwl, overflow = mo.groups()
                parsed_dict[key_str] = {"w_hpwl": w_hpwl, "overflow": overflow}

    return parsed_dict


def parse_log(
    placer_log_filename: str,
) -> Tuple[List[str], Dict[str, Dict[str, Dict[str, Any]]]]:
    """
    Returns
    - list of benchmark instance IDs
    - dictionary: benchmark instance ID -> process name -> key -> value
    """
    ins_count: Dict[str, int] = {}
    ins_list: List[str] = []
    ins_proc_dict: Dict[str, Dict[str, Dict[str, Any]]] = {}

    with open(placer_log_filename) as p_log:
        ins_key: str = ""
        for line in p_log:
            _dict = parse_a_line(line)
            if _dict:
                if "Input" in _dict:
                    if "path" in _dict["Input"]:
                        ins_path = _dict["Input"]["path"]
                        if ins_path not in ins_count:
                            ins_count[ins_path] = 1
                        else:
                            ins_count[ins_path] += 1
                        ins_key = ins_path + str(ins_count[ins_path])
                        ins_list.append(ins_key)
                if ins_key not in ins_proc_dict:
                    ins_proc_dict[ins_key] = {}
                for val_key, value in _dict.items():
                    if val_key not in ins_proc_dict[ins_key]:
                        ins_proc_dict[ins_key][val_key] = {}
                    ins_proc_dict[ins_key][val_key].update(value)

    return ins_list, ins_proc_dict


def write_to_xlsx(
    ins_list: List[str],
    ins_proc_dict: Dict[str, Dict[str, Dict[str, Any]]],
    output_filename: str,
):
    proc_header_seq = ["IP", "GP", "LG", "DP"]
    time_header_seq = ["Input"] + proc_header_seq + ["Output"]
    log_proc_dict = {
        "Input": "Input",
        "IP": "Initialplacement",
        "GP": "Globalplacement",
        "LG": "Legalization",
        "DP": "Detailedplacement",
        "Output": "Output",
    }
    wb = Workbook()

    # wHPWL sheet
    w_hpwl_ws: Worksheet = wb.active
    w_hpwl_ws.title = "wHPWL"
    # overflow sheet
    overflow_ws: Worksheet = wb.create_sheet("overflow")
    col_header = ["Instance"] + proc_header_seq
    w_hpwl_ws.append(col_header)
    overflow_ws.append(col_header)
    # runtime sheet
    runtime_ws: Worksheet = wb.create_sheet("runtime")
    col_header = ["Instance", "Total"] + time_header_seq
    runtime_ws.append(col_header)

    for ins_name in ins_list:
        wHPWL_dict: Dict[str, Any] = {}
        overflow_dict: Dict[str, Any] = {}
        for proc in proc_header_seq:
            log_proc = log_proc_dict[proc]
            val_dict = ins_proc_dict[ins_name][log_proc]
            wHPWL_dict[proc] = eval(val_dict["w_hpwl"])
            overflow_dict[proc] = eval(val_dict["overflow"])
        wHPWL_row = [ins_name] + [
            wHPWL_dict[proc_header] for proc_header in proc_header_seq
        ]
        w_hpwl_ws.append(wHPWL_row)
        overtime_row = [ins_name] + [
            overflow_dict[proc_header] for proc_header in proc_header_seq
        ]
        overflow_ws.append(overtime_row)

        runtime_dict: Dict[str, Any] = {"Total": 0.0}
        for proc in time_header_seq:
            log_proc = log_proc_dict[proc]
            val_dict = ins_proc_dict[ins_name][log_proc]
            runtime = eval(val_dict["runtime"])
            runtime_dict[proc] = runtime
            runtime_dict["Total"] += runtime
        runtime_row = [ins_name, runtime_dict["Total"]] + [
            runtime_dict[proc_header] for proc_header in time_header_seq
        ]
        runtime_ws.append(runtime_row)

    wb.save(filename=output_filename)


def main():
    root_logger = logging.getLogger()
    root_logger.setLevel(logging.INFO)
    # Logging to a .log file
    handler = logging.FileHandler("log_parser.log", encoding="utf-8")
    root_logger.addHandler(handler)
    # show log messages on terminal as well
    root_logger.addHandler(logging.StreamHandler(sys.stdout))

    global START_DT
    logging.info(f"{__name__} program start @ {START_DT}"[:-3])

    ins_list, ins_proc_dict = parse_log(PLACER_LOG_FILENAME)
    write_to_xlsx(ins_list, ins_proc_dict, OUTPUT_FILENAME)


if __name__ == "__main__":
    START_DT = datetime.datetime.now()
    main()
    END_DT = datetime.datetime.now()
    elapsed_d = END_DT - START_DT
    logging.info(
        f"{__name__} program end @ {END_DT}"[:-3] + f"; took total {elapsed_d}"[:-3]
    )
